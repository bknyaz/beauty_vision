# -*- coding: utf-8 -*-
"""
Created on Mon Aug 29 11:07:10 2016
@author: Boris Knyazev

This code is an attempt to reproduce results similar to
[1] Xie, D., Liang, L., Jin, L., Xu, J., & Li, M. (2015, October). 
    SCUT-FBP: A Benchmark Dataset for Facial Beauty Perception. 
    In Systems, Man, and Cybernetics (SMC), 2015 IEEE International Conference on (pp. 1821-1826). IEEE.

Disclaimer:
Measuring attractiveness or beauty is highly subjective, so this is just an experiment
In this dataset, I actually find many faces with low rating quite cute

1. Caffe for Python must be installed for this example
2. Caffe model must be downloaded from https://gist.github.com/ksimonyan/fd8800eeb36e276cd6f9#file-readme-md

"""

import caffe
import numpy as np
import matplotlib.pyplot as plt
import os
import io
import skimage.transform
import scipy

from openpyxl import load_workbook
from PIL import Image
from sklearn.decomposition import PCA, RandomizedPCA
from sklearn import svm
from sklearn.metrics import mean_absolute_error
from timeit import default_timer as timer
from theano.tensor.nnet import relu

from lasagne.utils import floatX

start = timer()

# options
SVR_kernel='linear' # or'rbf'
PCA_dim = 0 # the number of PCA components (0 - no PCA)

# load data
# Dataset can be downloaded from http://www.hcii-lab.net/data/SCUT-FBP/EN/download.html
data_dir = '/home/boris/Project/data/images/SCUT-FBP/';
print('loading data')
wb = load_workbook(filename=data_dir+'Rating_Collection/Attractiveness label.xlsx', read_only=True)
ws = wb['Sheet1']
image_files = os.listdir(data_dir+'Data_Collection/')
image_list = []
labels=[]
# we want to resize images, because it is not reasonable to work with very large images
face_sz = (224,294)
for r in range(ws.min_row+1, ws.max_row+1):
  labels.append((ws.cell('B%d' % r).value))
  image_path = data_dir+'Data_Collection' + '/SCUT-FBP-%d.jpg' % int(str(r-1))
  img = Image.open(open(image_path)).resize(face_sz)
  img = np.asarray(img, dtype='float64')[35:259,:,:] / 256. # crop [40:296,:,:]
  print(image_path + ', size: ' + str(img.shape) + ', attractiveness: %1.2f' % labels[r-2])
  image_list.append(img)

labels = np.asarray(labels)
print('%d images and %d labels read' % (len(image_list), len(labels)))

# plot distribution of ratings
hist, bin_edges = np.histogram(labels, density=True, bins=50)
fig = plt.figure()
ax = fig.add_subplot(111)
plt.scatter(bin_edges[0:len(hist)], hist)
ax.set_xlabel('Attractiveness rating')
ax.set_ylabel('Frequency')
ax.set_title('Distribution of ratings')
plt.show()

# show less, average and more beautiful faces (just to check if data makes sense)
#least_beautiful = np.argmin(labels)
#avg_beautiful = np.argmin(abs(labels-np.mean(labels)))
#most_beautiful = np.argmax(labels)
less_beautiful = np.random.permutation(np.nonzero(labels<2)[0])[0]
avg_beautiful = np.random.permutation(np.nonzero((labels>2) & (labels<3))[0])[0]
more_beautiful = np.random.permutation(np.nonzero(labels>4.5)[0])[0]
fig = plt.figure()
ax = fig.add_subplot(111)
ax.set_title('Attractiveness (low): %1.2f' % labels[less_beautiful])
plt.imshow(image_list[less_beautiful])
fig = plt.figure()
ax = fig.add_subplot(111)
ax.set_title('Attractiveness (avg): %1.2f' % labels[avg_beautiful])
plt.imshow(image_list[avg_beautiful])
fig = plt.figure()
ax = fig.add_subplot(111)
ax.set_title('Attractiveness (high): %1.2f' % labels[more_beautiful])
plt.imshow(image_list[more_beautiful])
plt.show()

print('computing features using the VGG-face pretrained network')

caffe.set_mode_gpu()
#Define the network
model_dir = '/home/boris/Project/3rd_party/models/vgg_cnn_s_caffe/'
net_caffe = caffe.Net(model_dir + 'VGG_CNN_S_deploy.prototxt', model_dir + 'VGG_CNN_S.caffemodel', caffe.TEST)

MEAN_RGB = np.array([129.1863,104.7624,93.5940]) # see vgg face matlab demo
#MEAN_RGB = scipy.io.loadmat(model_dir + '/VGG_mean.mat')['image_mean']

# process all images as a single batch (VGG_FACE_deploy.prototxt should be editted)
#image_list = np.stack(image_list)
#image_list = image_list*256
#image_list = image_list - MEAN_RGB
#image_list = np.swapaxes(np.swapaxes(image_list, 2, 3), 1, 2)
#image_list = image_list[:,::-1, :, :]
#out = net_caffe.forward(data = floatX(image_list), end='pool5')
#feat_maps = out['pool5'].reshape((out['pool5'].shape[0],np.prod(out['pool5'].shape[1:])))

# process images one by one
feat_maps = []    
for i, img in enumerate(image_list):
    im = np.copy(img*256)
    im = im - MEAN_RGB
    # convert to the caffe format
    im = np.swapaxes(np.swapaxes(im, 1, 2), 0, 1) # -> channels x height x width
    im = im[::-1, :, :] # RGB -> BGR
    im = floatX(im[np.newaxis]) # add axis -> 1 x channels x height x width
    out1 = net_caffe.forward(data = im, end='pool5')
#    feat_maps.append(out1['pool5'].flatten())
    out2 = net_caffe.forward(data = im, end='fc6')
    feat_maps.append(np.concatenate((out1['pool5'].flatten(),2*relu(out2['fc6'].flatten()))))
    print('img %d' % i)
    
feat_maps = np.asarray(feat_maps)

if PCA_dim > 0:
    pca = RandomizedPCA(n_components=PCA_dim) 

# 10-fold cross-validation
print('performing cross-validation using SVM regression')
n_folds = 10

# shuffle samples
ids = np.random.permutation(feat_maps.shape[0])
feat_maps = feat_maps[ids,:]
labels_ = labels[ids]
n = len(labels_)/n_folds

PC = []
MAE = []
for fold_id in range(n_folds):
    test_ids = np.arange(fold_id*n,(fold_id+1)*n)
    train_ids = np.concatenate((np.arange(0,fold_id*n),
    np.arange((fold_id+1)*n,len(labels_))))
    assert(len(train_ids) == n*(n_folds-1))
    assert(len(test_ids) == n)
    feat_maps_train = feat_maps[train_ids,:]
    feat_maps_test = feat_maps[test_ids,:]
    if PCA_dim > 0:
        feat_maps_train = pca.fit_transform(feat_maps_train)
        feat_maps_test = pca.transform(feat_maps_test)
        feat_maps_train = (feat_maps_train - np.mean(feat_maps_train,0))/np.std(feat_maps_train,0)
        feat_maps_test = (feat_maps_test - np.mean(feat_maps_test,0))/np.std(feat_maps_test,0)
    else:
        feat_maps_train = (feat_maps_train - np.mean(feat_maps_train,0))
        feat_maps_test = (feat_maps_test - np.mean(feat_maps_test,0))

    clf = svm.SVR(kernel=SVR_kernel, C=1)
    clf.fit(feat_maps_train, labels_[train_ids])
    pred = clf.predict(feat_maps_test)
    PC.append(np.corrcoef(labels_[test_ids],pred)[0,1])
    MAE.append(mean_absolute_error(labels_[test_ids],pred))
    
    fig = plt.figure()
    plt.plot(np.arange(n),labels_[test_ids], label='true')
    plt.plot(np.arange(n),pred, label='predicted')
    plt.legend(loc='upper left')
    plt.title('fold %d, PC = %1.2f, MAE = %1.2f' % 
    (fold_id,np.corrcoef(labels_[test_ids],pred)[0,1],mean_absolute_error(labels_[test_ids],pred)))
    plt.show()

print('PC (Pearson correlation) mean = %1.2f (random guess = %1.2f)' % (np.mean(PC),np.corrcoef(labels_[test_ids],2.5*np.ones((len(test_ids),))+0.1*np.random.rand(len(test_ids),)
)[0,1]))
print('MAE (Mean absolute error) mean = %1.2f (random guess = %1.2f)' % (np.mean(MAE),mean_absolute_error(labels_[test_ids],2.5*np.ones((len(test_ids),)))))

end = timer()
print('Test took %1.2f sec' % (end - start))