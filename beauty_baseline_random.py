# -*- coding: utf-8 -*-
"""
Created on Tue Aug 23 18:54:47 2016
@author: Boris Knyazev

This code is an attempt to reproduce results similar to
[1] Xie, D., Liang, L., Jin, L., Xu, J., & Li, M. (2015, October). 
    SCUT-FBP: A Benchmark Dataset for Facial Beauty Perception. 
    In Systems, Man, and Cybernetics (SMC), 2015 IEEE International Conference on (pp. 1821-1826). IEEE.

Disclaimer:
Measuring attractiveness or beauty is highly subjective, so this is just an experiment
In this dataset, I actually find many faces with low rating quite cute

"""

from __future__ import print_function

import numpy as np
import matplotlib.pyplot as plt
import os
import theano

from openpyxl import load_workbook
from PIL import Image
from sklearn.decomposition import PCA
from sklearn import svm
from sklearn.metrics import mean_absolute_error

from theano import tensor as T
from theano.tensor.nnet import conv2d
from theano.tensor.signal import pool
from theano.tensor.nnet import relu

# load data
# Dataset can be downloaded from http://www.hcii-lab.net/data/SCUT-FBP/EN/download.html
data_dir = '/home/boris/Project/data/images/SCUT-FBP/';
print('loading data')
wb = load_workbook(filename=data_dir+'Rating_Collection/Attractiveness label.xlsx', read_only=True)
ws = wb['Sheet1']
image_files = os.listdir(data_dir+'Data_Collection/')
image_list = []
labels=[]
# we want to resize images, because it is not reasonable to work with very large images
face_sz = (256,336) # this size will keep aspect ratio for most of the images
for r in range(ws.min_row+1, ws.max_row+1):
  labels.append((ws.cell('B%d' % r).value))
  image_path = data_dir+'Data_Collection' + '/SCUT-FBP-%d.jpg' % int(str(r-1))
  img = Image.open(open(image_path)).resize(face_sz)
  img = np.asarray(img, dtype='float64') / 256.
  print(image_path + ', size: ' + str(img.shape) + ', attractiveness: %1.2f' % labels[r-2])
  image_list.append(img)

print('%d images and %d labels read' % (len(image_list), len(labels)))

# plot distribution of ratings
hist, bin_edges = np.histogram(labels, density=True, bins=50)
fig = plt.figure()
ax = fig.add_subplot(111)
plt.scatter(bin_edges[0:len(hist)], hist)
ax.set_xlabel('Attractiveness rating')
ax.set_ylabel('Frequency')

# show least, average and most beautiful faces (just to check if data makes sense)
least_beautiful = np.argmin(labels)
avg_beautiful = np.argmin(abs(labels-np.mean(labels)))
most_beautiful = np.argmax(labels)
fig = plt.figure()
ax = fig.add_subplot(111)
ax.set_title('Attractiveness (least): %1.2f' % labels[least_beautiful])
plt.imshow(image_list[least_beautiful])
fig = plt.figure()
ax = fig.add_subplot(111)
ax.set_title('Attractiveness (avg): %1.2f' % labels[avg_beautiful])
plt.imshow(image_list[avg_beautiful])
fig = plt.figure()
ax = fig.add_subplot(111)
ax.set_title('Attractiveness (max): %1.2f' % labels[most_beautiful])
plt.imshow(image_list[most_beautiful])


print('computing features using random filters')
# instantiate 4D tensor for input
# from http://deeplearning.net/tutorial/lenet.html
input = T.tensor4(name='input')

# initialize shared variable for weights.
# 16 random filters
rng = np.random.RandomState(23455)
w_shp = (16, 3, 9, 9)
w_bound = np.sqrt(3 * 9 * 9)
W_arr = np.asarray(rng.uniform(low=-1.0 / w_bound,
                high=1.0 / w_bound,
                size=w_shp),
            dtype=input.dtype)
W = theano.shared(W_arr, name ='W')

# convolution, max-pooling and ReLU
f1 = theano.function([input], conv2d(input, W))
f2 = theano.function([input], pool.pool_2d(input, (16, 16), ignore_border=True))
feat_maps = relu(f2(f1(np.stack(image_list).transpose(0,3,1,2))))
feat_maps = feat_maps.reshape(feat_maps.shape[0],np.prod(feat_maps.shape[1:]))

pca = PCA(n_components=50) # this is just a heuristic choice

# 10-fold cross-validation
print('performing cross-validation using SVM regression')
n_folds = 10
ids = np.random.permutation(feat_maps.shape[0])
feat_maps = feat_maps[ids,:]
labels_ = np.asarray(labels)[ids]
n = len(labels_)/n_folds

PC = []
MAE = []
for fold_id in range(n_folds):
    test_ids = np.arange(fold_id*n,(fold_id+1)*n)
    train_ids = np.concatenate((np.arange(0,fold_id*n),
    np.arange((fold_id+1)*n,len(labels_))))
    assert(len(train_ids) == 450)
    assert(len(test_ids) == n)
    feat_maps_train = pca.fit_transform(feat_maps[train_ids,:])
    feat_maps_test = pca.transform(feat_maps[test_ids,:])
    
    clf = svm.SVR()
    clf.fit(feat_maps_train, labels_[train_ids])
    pred = clf.predict(feat_maps_test)
    PC.append(np.corrcoef(labels_[test_ids],pred)[0,1])
    MAE.append(mean_absolute_error(labels_[test_ids],pred))
    
    fig = plt.figure()
    plt.plot(np.arange(n),labels_[test_ids], label='true')
    plt.plot(np.arange(n),pred, label='predicted')
    plt.legend(loc='upper left')
    plt.title('fold %d, PC = %1.2f, MAE = %1.2f' % 
    (fold_id,np.corrcoef(labels_[test_ids],pred)[0,1],mean_absolute_error(labels_[test_ids],pred)))
    plt.show()

print('PC (Pearson correlation) mean = %1.2f' % np.mean(PC))
print('MAE (Mean absolute error) mean = %1.2f' % np.mean(MAE))